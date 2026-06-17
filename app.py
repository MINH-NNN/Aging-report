#!/usr/bin/env python3
"""FastAPI wrapper cho Aging Report Agent — output: 4 combined files."""

import os, sys, json, re, shutil, subprocess, tempfile
from datetime import datetime
from pathlib import Path
from typing import List
from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
import uvicorn

app = FastAPI(title="Aging Report Agent", version="2.0.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

SCRIPTS  = Path(__file__).parent / "scripts"
BASE_DIR = Path(__file__).parent

# ── Đọc .env ──────────────────────────────────────────────────────────────────
def _load_env():
    env_path = BASE_DIR / ".env"
    if env_path.exists():
        for line in env_path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                os.environ.setdefault(k.strip(), v.strip())

_load_env()

# ── Tracker (lưu lịch sử nhắc nợ) ────────────────────────────────────────────
_last_output_dir: str = ""
_dashboard_html: str = ""
_last_ar_by_pic: dict = {}   # lưu by_pic sau khi parse AR, dùng cho /send
_last_report_date: str = "N/A"
TRACKER_PATH = BASE_DIR / "debt_tracker.json"

# Mapping PIC → email, load từ pic_emails.json, có thể update qua /api/upload-contacts
def _load_pic_emails() -> dict:
    p = BASE_DIR / "pic_emails.json"
    if p.exists():
        try: return json.loads(p.read_text(encoding="utf-8"))
        except: pass
    return {}
_pic_emails: dict = _load_pic_emails()

def load_tracker() -> dict:
    if TRACKER_PATH.exists():
        try: return json.loads(TRACKER_PATH.read_text(encoding="utf-8"))
        except: pass
    return {}

def save_tracker(tracker: dict):
    TRACKER_PATH.write_text(json.dumps(tracker, ensure_ascii=False, indent=2), encoding="utf-8")

def tracker_key(row: dict) -> str:
    inv = str(row.get("invoice") or "").strip()
    ent = str(row.get("entity") or "").strip()
    return f"{inv}|{ent}" if inv else f"{ent}|{row.get('inv_date','')}"

def ar_tracker_key(r: dict) -> str:
    """Key từ AR record (invoice_no/code thay vì invoice/entity)."""
    inv = str(r.get("invoice_no") or "").strip()
    ent = str(r.get("code") or "").strip()
    return f"{inv}|{ent}" if inv else f"{ent}|{r.get('invoice_date','')}"

def merge_tracker(rows: list) -> list:
    """Merge tracker vào ROWS: restore email_sent, pic_resp, ghi_chu, reminder_count."""
    tracker = load_tracker()
    for row in rows:
        key = tracker_key(row)
        rec = tracker.get(key, {})
        row["reminder_count"] = rec.get("reminder_count", 0)
        row["email_sent"]     = rec.get("email_sent", "")
        row["pic_resp"]       = rec.get("pic_resp", row.get("pic_resp", "Not Reminded"))
        row["ghi_chu"]        = rec.get("ghi_chu", "")
    return rows

def update_tracker(updates: dict):
    """Merge updates {key: {field: value}} vào tracker và lưu."""
    tracker = load_tracker()
    for key, fields in updates.items():
        if key not in tracker:
            tracker[key] = {}
        tracker[key].update(fields)
    save_tracker(tracker)

# ── Dashboard ─────────────────────────────────────────────────────────────────
import sys as _sys
_sys.path.insert(0, str(BASE_DIR / "scripts"))
from generate_combined_dashboard import build_html as _build_html
from datetime import datetime as _dt

def _empty_dashboard():
    today = _dt.now().strftime("%d/%m/%Y %H:%M")
    return _build_html("[]", "Chua co du lieu", today)


# ── Helper chạy script con ────────────────────────────────────────────────────
def run_script(script_name, args):
    env = os.environ.copy()
    env["PYTHONIOENCODING"] = "utf-8"
    result = subprocess.run(
        [sys.executable, str(SCRIPTS / script_name)] + args,
        capture_output=True, text=True, encoding="utf-8", env=env
    )
    if result.returncode != 0:
        raise HTTPException(status_code=500,
            detail=f"[{script_name}] {result.stderr.strip() or result.stdout.strip()}")
    return result.stdout


# ── Email draft helpers ───────────────────────────────────────────────────────
def _generate_ar_email_lines(data: dict) -> list:
    report_date = data.get("report_date", "")
    by_pic = data.get("by_pic", {})
    NO_PIC_KEYS = {"Chua phan cong", "Chưa phân công"}
    pic_records = {k: v for k, v in by_pic.items() if k not in NO_PIC_KEYS}
    no_pic_records = next((v for k, v in by_pic.items() if k in NO_PIC_KEYS), [])
    lines = [f"\n---\n## AR — {report_date}\n"]
    AGING = {
        "Over 180 ngày": lambda d: f"Quá hạn {d} ngày 🔴🔴",
        "91-180 ngày":   lambda d: f"Quá hạn {d} ngày 🔴",
        "61-90 ngày":    lambda d: f"Quá hạn {d} ngày ⚠️",
        "31-60 ngày":    lambda d: f"Quá hạn {d} ngày",
        "1-30 ngày":     lambda d: f"Quá hạn {d} ngày",
        "Current":       lambda d: "Chưa đến hạn",
        "N/A":           lambda d: "N/A",
    }
    for pic, records in pic_records.items():
        rows = ""
        for i, r in enumerate(records, 1):
            fn   = AGING.get(r["aging_bucket"], lambda d: r["aging_bucket"])
            name = f"{r['code']} - {r['name']}" if r.get("name") else str(r.get("code",""))
            rows += f"| {i} | {name} | {r.get('invoice_no') or 'N/A'} | {r.get('invoice_date') or 'N/A'} | {r.get('base_amount',0):,.0f} | {fn(r.get('over_day',0))} | {(r.get('description') or '').strip()} |\n"
        table = "| # | Khách hàng | Invoice | Ngày HĐ | Số tiền | Tình trạng | Mô tả |\n|---|---|---|---|---|---|---|\n" + rows
        lines.append(f"---\n\n## PIC: {pic}\n\n**Subject:** [AR] Thông báo công nợ — {report_date}\n\n{table}\nVui lòng đôn đốc khách hàng thanh toán.\n\nTrân trọng,\n[Ký tên]\n")
    if no_pic_records:
        by_cust: dict = {}
        for r in no_pic_records:
            key = (r.get("code",""), r.get("name") or str(r.get("code","")))
            by_cust.setdefault(key, []).append(r)
        for (code, cname), recs in by_cust.items():
            total = sum(r.get("base_amount",0) for r in recs)
            rows = ""
            for r in recs:
                rows += f"| {cname} | {(r.get('description') or '').strip()} | {r.get('invoice_date') or 'N/A'} | {r.get('base_amount',0):,.0f} |\n"
            table = "| Customer Name | Description | Invoice Date | Amount |\n|---|---|---|---|\n" + rows + f"| **Tổng Cộng** | | | **{total:,.0f}** |\n"
            lines.append(f"---\n\n## {cname}\n\n**Subject:** [AR] THƯ NHẮC NỢ THANH TOÁN — {report_date}\n\n{table}\nVui lòng kiểm tra và thanh toán.\n\nTrân trọng,\n[Ký tên]\n")
    return lines


def _generate_new_email_lines(data: dict, period: str) -> list:
    account = data.get("account", "")
    lines = [f"\n---\n## TK{account} — Kỳ {period}\n"]
    if data.get("by_employee"):
        for emp, records in data["by_employee"].items():
            total = sum(r.get("amount",0) for r in records)
            rows = "".join(f"| {r.get('invoice_date') or 'N/A'} | {r.get('invoice_no','')} | {r.get('description','')} | {r.get('amount',0):,.0f} | {r.get('reimbursement_date') or 'Chưa có'} |\n" for r in records)
            table = "| Ngày CT | Số CT | Diễn giải | Số tiền (VND) | Ngày hoàn ứng |\n|---|---|---|---|---|\n" + rows + f"| | | **Tổng cộng** | **{total:,.0f}** | |\n"
            lines.append(f"---\n\n## {emp}\n\n**Subject:** [Nhắc hoàn ứng] Kỳ {period}\n\n{table}\nVui lòng hoàn tất hồ sơ.\n\nTrân trọng,\nBộ phận Kế toán\n")
    elif data.get("by_vendor"):
        for vendor, records in data["by_vendor"].items():
            total = sum(r.get("total_vnd",0) for r in records)
            rows = "".join(f"| {r.get('invoice_date') or 'N/A'} | {r.get('invoice_no','')} | {r.get('description','')} | {r.get('total_vnd',0):,.0f} |\n" for r in records)
            table = "| Ngày HĐ | Số HĐ | Diễn giải | Tổng VND |\n|---|---|---|---|\n" + rows + f"| | | **Tổng cộng** | **{total:,.0f}** |\n"
            lines.append(f"---\n\n## {vendor}\n\n**Subject:** [AP TK{account}] Công nợ kỳ {period}\n\n{table}\n\nTrân trọng,\nBộ phận Kế toán\n")
    elif data.get("by_supplier"):
        for sup, records in data["by_supplier"].items():
            total = sum(r.get("amount",0) for r in records)
            rows = "".join(f"| {r.get('gl_date') or 'N/A'} | {r.get('invoice_no','')} | {r.get('description','')} | {r.get('amount',0):,.0f} |\n" for r in records)
            table = "| Ngày GL | Số HĐ | Diễn giải | Số tiền |\n|---|---|---|---|\n" + rows + f"| | | **Tổng cộng** | **{total:,.0f}** |\n"
            lines.append(f"---\n\n## {sup}\n\n**Subject:** [TK{account}] Nhắc công nợ kỳ {period}\n\n{table}\n\nTrân trọng,\nBộ phận Kế toán\n")
    return lines


# ── Routes ────────────────────────────────────────────────────────────────────
@app.get("/health")
def health():
    return {"status": "ok"}


@app.get("/", response_class=HTMLResponse)
def ui():
    return _dashboard_html if _dashboard_html else _empty_dashboard()


@app.post("/process")
async def process(files: List[UploadFile] = File(...)):
    """Upload 1+ file Excel, trả về ZIP gồm Master, SLA Tracker, Dashboard, Email Drafts."""
    global _last_output_dir, _dashboard_html
    work_dir = tempfile.mkdtemp()
    try:
        output_dir = Path(work_dir) / "output"
        output_dir.mkdir()
        bundle_period = datetime.now().strftime("T%m_%Y")
        all_parsed = []

        for upload in files:
            safe_name = re.sub(r'[^\w_. -]', '_', upload.filename)
            file_path = os.path.join(work_dir, safe_name)
            with open(file_path, "wb") as f:
                f.write(await upload.read())

            type_out  = run_script("detect_type.py", [file_path])
            type_info = json.loads(type_out)
            ftype   = type_info.get("type", "UNKNOWN")
            period  = type_info.get("period", bundle_period)
            account = type_info.get("account", "")

            if period in ("N/A", None, ""):
                period = bundle_period
            else:
                period = "T" + str(period).replace("/", "_")

            if ftype == "AR":
                parsed_data = json.loads(run_script("parse_ar.py", [file_path]))
                parsed_data.update({"file_type":"AR","period":period,"source_file":upload.filename,"account":"AR"})
                all_parsed.append(parsed_data)
            elif ftype in ("TAM_UNG", "AP_AGING", "PREPAY"):
                parser_map = {"TAM_UNG":"parse_tam_ung.py","AP_AGING":"parse_ap_aging.py","PREPAY":"parse_prepay.py"}
                parsed_data = json.loads(run_script(parser_map[ftype], [file_path]))
                parsed_data.update({"file_type":ftype,"period":period,"source_file":upload.filename})
                all_parsed.append(parsed_data)

        if not all_parsed:
            raise HTTPException(status_code=422, detail="Không nhận dạng được loại file nào.")

        combined_path = os.path.join(work_dir, "parsed_combined.json")
        Path(combined_path).write_text(
            json.dumps({"file_type":"COMBINED","period":bundle_period,"sources":all_parsed}, ensure_ascii=False),
            encoding="utf-8")

        run_script("generate_combined_master.py",      [combined_path, str(output_dir / f"Master_{bundle_period}.xlsx")])
        run_script("generate_combined_sla_tracker.py", [combined_path, str(output_dir / f"SLA_Tracker_{bundle_period}.xlsx")])
        run_script("generate_combined_dashboard.py",   [combined_path, str(output_dir / f"Dashboard_{bundle_period}.html")])

        email_lines = [f"# Email Drafts — {bundle_period}\n"]
        for src in all_parsed:
            if src.get("file_type") == "AR":
                email_lines.extend(_generate_ar_email_lines(src))
            else:
                email_lines.extend(_generate_new_email_lines(src, src.get("period", bundle_period)))
        Path(str(output_dir / f"EmailDrafts_{bundle_period}.md")).write_text("\n".join(email_lines), encoding="utf-8")

        _last_output_dir = str(output_dir)
        dash_path = str(output_dir / f"Dashboard_{bundle_period}.html")
        _dashboard_html = open(dash_path, encoding="utf-8").read()

        zip_base = os.path.join(work_dir, "ZionReports")
        shutil.make_archive(zip_base, "zip", str(output_dir))
        return FileResponse(path=zip_base+".zip", media_type="application/zip", filename="ZionReports_output.zip")

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/process")
async def api_process(files: List[UploadFile] = File(...)):
    """Giống /process nhưng trả về JSON rows để cập nhật dashboard trực tiếp."""
    global _last_output_dir, _dashboard_html, _last_ar_by_pic
    work_dir = tempfile.mkdtemp()
    try:
        output_dir = Path(work_dir) / "output"
        output_dir.mkdir()
        bundle_period = datetime.now().strftime("T%m_%Y")
        all_parsed = []

        for upload in files:
            safe_name = re.sub(r'[^\w_. -]', '_', upload.filename)
            file_path = os.path.join(work_dir, safe_name)
            with open(file_path, "wb") as f:
                f.write(await upload.read())

            type_info = json.loads(run_script("detect_type.py", [file_path]))
            ftype   = type_info.get("type", "UNKNOWN")
            period  = type_info.get("period", bundle_period)
            if period in ("N/A", None, ""):
                period = bundle_period
            else:
                period = "T" + str(period).replace("/", "_")

            if ftype == "AR":
                parsed_data = json.loads(run_script("parse_ar.py", [file_path]))
                parsed_data.update({"file_type":"AR","period":period,"source_file":upload.filename,"account":"AR"})
                all_parsed.append(parsed_data)
            elif ftype in ("TAM_UNG", "AP_AGING", "PREPAY"):
                parser_map = {"TAM_UNG":"parse_tam_ung.py","AP_AGING":"parse_ap_aging.py","PREPAY":"parse_prepay.py"}
                parsed_data = json.loads(run_script(parser_map[ftype], [file_path]))
                parsed_data.update({"file_type":ftype,"period":period,"source_file":upload.filename})
                all_parsed.append(parsed_data)

        if not all_parsed:
            raise HTTPException(status_code=422, detail="Khong nhan dang duoc loai file nao.")

        # Lưu by_pic và report_date từ AR file cuối cùng được parse
        for src in all_parsed:
            if src.get("file_type") == "AR":
                _last_ar_by_pic = src.get("by_pic", {})
                _last_report_date = src.get("report_date", "N/A")

        combined_path = os.path.join(work_dir, "parsed_combined.json")
        Path(combined_path).write_text(
            json.dumps({"file_type":"COMBINED","period":bundle_period,"sources":all_parsed}, ensure_ascii=False),
            encoding="utf-8")

        try:
            run_script("generate_combined_master.py",      [combined_path, str(output_dir / f"Master_{bundle_period}.xlsx")])
            run_script("generate_combined_sla_tracker.py", [combined_path, str(output_dir / f"SLA_Tracker_{bundle_period}.xlsx")])
            run_script("generate_combined_dashboard.py",   [combined_path, str(output_dir / f"Dashboard_{bundle_period}.html")])
            email_lines = [f"# Email Drafts — {bundle_period}\n"]
            for src in all_parsed:
                if src.get("file_type") == "AR":
                    email_lines.extend(_generate_ar_email_lines(src))
                else:
                    email_lines.extend(_generate_new_email_lines(src, src.get("period", bundle_period)))
            Path(str(output_dir / f"EmailDrafts_{bundle_period}.md")).write_text("\n".join(email_lines), encoding="utf-8")
            _last_output_dir = str(output_dir)
            _dashboard_html = open(str(output_dir / f"Dashboard_{bundle_period}.html"), encoding="utf-8").read()
        except Exception:
            pass

        rows_data = json.loads(run_script("generate_combined_dashboard.py", [combined_path, "--json"]))
        rows_data["rows"] = merge_tracker(rows_data["rows"])

        # Rebuild _dashboard_html với tracker data để reload trang vẫn giữ state
        try:
            _dashboard_html = _build_html(
                json.dumps(rows_data["rows"], ensure_ascii=False),
                rows_data.get("report_date", ""),
                rows_data.get("today", _dt.now().strftime("%d/%m/%Y"))
            )
        except Exception:
            pass

        return JSONResponse(rows_data)

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/download")
def download():
    if not _last_output_dir:
        raise HTTPException(status_code=404, detail="Chua co du lieu. Upload file truoc.")
    zip_base = os.path.join(tempfile.mkdtemp(), "ZionReports")
    shutil.make_archive(zip_base, "zip", _last_output_dir)
    return FileResponse(path=zip_base+".zip", media_type="application/zip", filename="ZionReports_output.zip")


@app.post("/api/clear")
def api_clear():
    global _dashboard_html
    _dashboard_html = ""
    return {"status": "cleared"}


@app.post("/api/upload-contacts")
async def upload_contacts(file: UploadFile = File(...)):
    """Upload file Mail Contact (.xlsx) để cập nhật mapping PIC → email cho session hiện tại."""
    global _pic_emails
    import openpyxl, io, csv as csv_mod
    raw = await file.read()
    fname = file.filename.lower()

    rows_raw = []
    if fname.endswith(".csv"):
        text = raw.decode("utf-8-sig", errors="replace")
        rows_raw = list(csv_mod.reader(text.splitlines()))
    else:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.active
        rows_raw = [[str(c).strip() if c is not None else "" for c in row]
                    for row in ws.iter_rows(values_only=True)]

    # Tìm header row có cột email/mail
    header_idx = None
    headers = []
    for i, row in enumerate(rows_raw):
        if any("mail" in str(v).lower() or "email" in str(v).lower() for v in row if v):
            header_idx = i
            headers = [str(v).strip() for v in row]
            break

    if header_idx is None:
        raise HTTPException(status_code=422, detail="Không tìm thấy cột Email trong file.")

    def find_col(*candidates):
        for c in candidates:
            for i, h in enumerate(headers):
                if c.lower() in h.lower():
                    return i
        return None

    pic_idx   = find_col("pic", "name", "ten", "contact")
    email_idx = find_col("email", "mail")

    if email_idx is None:
        raise HTTPException(status_code=422, detail="Không tìm thấy cột Email.")

    mapping = {}
    for row in rows_raw[header_idx + 1:]:
        if len(row) <= email_idx:
            continue
        email = str(row[email_idx]).strip()
        if not email or "@" not in email:
            continue
        if pic_idx is not None and pic_idx < len(row) and row[pic_idx]:
            key = str(row[pic_idx]).strip()
        else:
            key = email.split("@")[0]
        if key:
            mapping[key] = email

    _pic_emails.update(mapping)
    return {"status": "ok", "updated": len(mapping), "mapping": mapping}


@app.get("/api/pic-template")
def pic_template(pics: str = "[]"):
    """Tạo file xlsx template cho danh sách PIC chưa có email."""
    import openpyxl, io
    try:
        pic_list = json.loads(pics)
    except Exception:
        pic_list = []
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mail Contact"
    ws.append(["PIC", "Email"])
    for p in pic_list:
        ws.append([p, ""])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=mail_contact_template.xlsx"}
    )


@app.get("/api/pics")
def get_pics():
    """Trả về danh sách PIC có khoản overdue (bỏ Current) + trạng thái email."""
    if not _last_ar_by_pic:
        raise HTTPException(status_code=404, detail="Chưa có dữ liệu AR. Upload file trước.")
    result = []
    for pic, records in _last_ar_by_pic.items():
        overdue = [r for r in records if r.get("aging_bucket") != "Current"]
        if not overdue:
            continue
        result.append({
            "pic":    pic,
            "email":  _pic_emails.get(pic),
            "count":  len(overdue),
            "mapped": pic in _pic_emails,
        })
    return {"pics": result}


@app.post("/send")
async def send_emails_endpoint(request: dict = None):
    """Gửi email cho các PIC được chọn. Body: {"pics": ["MinhNNN", ...], "dry_run": false}"""
    from fastapi import Request
    raise HTTPException(status_code=405, detail="Dùng /api/send thay thế")


@app.post("/api/send")
async def api_send(request_body: dict):
    """Gửi email cho các PIC được chọn.
    Body JSON: {"pics": ["MinhNNN", "QuyenNPT"], "dry_run": false}
    """
    global _last_report_date
    if not _last_ar_by_pic:
        raise HTTPException(status_code=404, detail="Chưa có dữ liệu AR. Upload file trước.")

    selected_pics = request_body.get("pics", list(_last_ar_by_pic.keys()))
    dry_run = request_body.get("dry_run", False)
    pic_emails = _pic_emails

    invoice_reminders = request_body.get("invoice_reminders", {})

    # Lọc by_pic theo danh sách được chọn, chỉ giữ khoản overdue (bỏ Current)
    filtered_by_pic = {}
    for p, v in _last_ar_by_pic.items():
        if p not in selected_pics:
            continue
        records = []
        for r in v:
            if r.get("aging_bucket") == "Current":
                continue
            rec = dict(r)
            inv_key = str(rec.get("invoice_no") or "") or f"{rec.get('code','')}{rec.get('invoice_date','')}"
            rec["reminder_count"] = invoice_reminders.get(inv_key, 1)
            records.append(rec)
        if records:
            filtered_by_pic[p] = records

    work_dir = tempfile.mkdtemp()
    try:
        # Ghi tạm filtered data + pic_emails để send_emails.py đọc
        filtered_path = os.path.join(work_dir, "filtered_by_pic.json")
        filtered_emails_path = os.path.join(work_dir, "pic_emails.json")
        Path(filtered_path).write_text(json.dumps({"by_pic": filtered_by_pic, "report_date": _last_report_date}, ensure_ascii=False), encoding="utf-8")
        Path(filtered_emails_path).write_text(json.dumps(pic_emails, ensure_ascii=False), encoding="utf-8")

        args = [filtered_path, filtered_emails_path, "--from-json"]
        if dry_run:
            args.append("--dry-run")

        env = os.environ.copy()
        env["PYTHONIOENCODING"] = "utf-8"
        result = subprocess.run(
            [sys.executable, str(SCRIPTS / "send_emails.py")] + args,
            capture_output=True, text=True, encoding="utf-8", env=env
        )
        if result.returncode != 0:
            raise HTTPException(status_code=500, detail=result.stderr.strip() or result.stdout.strip())

        # Lưu email_sent + reminder_count vào tracker
        if not dry_run:
            now_str = datetime.now().strftime("%H:%M %d/%m/%Y")
            tracker_updates = {}
            for pic, records in filtered_by_pic.items():
                for r in records:
                    key = ar_tracker_key(r)
                    old = load_tracker().get(key, {})
                    tracker_updates[key] = {
                        "entity":         str(r.get("code") or ""),
                        "invoice":        str(r.get("invoice_no") or ""),
                        "email_sent":     now_str,
                        "reminder_count": old.get("reminder_count", 0) + 1,
                        "pic_resp":       "Awaiting",
                        "ghi_chu":        old.get("ghi_chu", ""),
                    }
            if tracker_updates:
                update_tracker(tracker_updates)

        return {"status": "ok", "detail": result.stdout.strip()}
    finally:
        shutil.rmtree(work_dir, ignore_errors=True)


GSHEET_CSV = "https://docs.google.com/spreadsheets/d/1k8dALKES9y70SGWlTGDwXR_lo5gPNE6g76qRagCJjEA/export?format=csv"

@app.get("/api/sync-responses")
def sync_responses():
    """Fetch Google Form responses tu Google Sheet, tra ve mapping PIC name -> response data."""
    import urllib.request as _req
    import csv as _csv
    import unicodedata

    def norm(s: str) -> str:
        """Chuẩn hóa: bỏ dấu, lower, bỏ khoảng trắng thừa — để match header tiếng Việt."""
        s = unicodedata.normalize("NFD", s)
        s = "".join(c for c in s if unicodedata.category(c) != "Mn")
        return s.lower().strip()

    try:
        request = _req.Request(GSHEET_CSV, headers={"User-Agent": "Mozilla/5.0"})
        with _req.urlopen(request, timeout=15) as r:
            raw = r.read().decode("utf-8-sig", errors="replace")
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Khong tai duoc Google Sheet: {e}")

    rows = list(_csv.reader(raw.splitlines()))
    if not rows:
        return {"responses": {}, "total": 0}

    headers = [str(v).strip() for v in rows[0]]
    headers_norm = [norm(h) for h in headers]

    def find_col(*keywords):
        for kw in keywords:
            kw_n = norm(kw)
            for i, h in enumerate(headers_norm):
                if kw_n in h:
                    return i
        return None

    name_idx      = find_col("ho va ten", "ho ten", "name", "ten")
    nguyen_idx    = find_col("nguyen nhan", "ly do")
    ngay_idx      = find_col("ngay du kien", "ngay thanh toan")
    sotien_idx    = find_col("so tien", "tien du kien")
    ghichu_idx    = find_col("ghi chu khac", "ghi chu")
    timestamp_idx = find_col("timestamp", "thoi gian")

    if name_idx is None:
        # Debug: trả về headers để dễ kiểm tra
        raise HTTPException(status_code=422, detail=f"Khong tim thay cot 'Ho va ten'. Headers: {headers}")

    def get_val(row, idx):
        return str(row[idx]).strip() if idx is not None and idx < len(row) else ""

    result = {}
    for row in rows[1:]:
        if not any(row):
            continue
        name = get_val(row, name_idx)
        if not name:
            continue
        # Dùng name gốc làm key để match với row.pic trong dashboard
        result[name] = {
            "timestamp":    get_val(row, timestamp_idx),
            "nguyen_nhan":  get_val(row, nguyen_idx),
            "ngay_du_kien": get_val(row, ngay_idx),
            "so_tien":      get_val(row, sotien_idx),
            "ghi_chu":      get_val(row, ghichu_idx),
        }

    # Lưu "Responded" + ghi_chu vào tracker cho các invoice đã được gửi mail
    import unicodedata as _ud
    def _norm(s):
        s = _ud.normalize("NFD", s)
        return "".join(c for c in s if _ud.category(c) != "Mn").lower().strip()

    tracker_updates = {}
    for pic_name, records in _last_ar_by_pic.items():
        resp_data = result.get(pic_name) or next(
            (v for k, v in result.items() if _norm(k) == _norm(pic_name)), None)
        if not resp_data:
            continue
        for r in records:
            key = ar_tracker_key(r)
            existing = load_tracker().get(key, {})
            if not existing.get("email_sent"):
                continue  # chỉ update nếu đã gửi mail
            notes = []
            if resp_data.get("nguyen_nhan"): notes.append(resp_data["nguyen_nhan"])
            if resp_data.get("ngay_du_kien"): notes.append(f"Dự kiến: {resp_data['ngay_du_kien']}")
            if resp_data.get("so_tien"): notes.append(f"Số tiền: {resp_data['so_tien']}")
            if resp_data.get("ghi_chu"): notes.append(resp_data["ghi_chu"])
            tracker_updates[key] = {
                **existing,
                "pic_resp": "Responded",
                "ghi_chu":  " | ".join(notes) if notes else existing.get("ghi_chu", ""),
            }
    if tracker_updates:
        update_tracker(tracker_updates)

    return {"responses": result, "total": len(result), "headers": headers}


@app.post("/api/tracker/update")
async def api_tracker_update(request_body: dict):
    """Dashboard push state changes (No Response, ghi_chu edits) lên server."""
    updates = request_body.get("updates", {})
    if updates:
        update_tracker(updates)
    return {"ok": True}


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8080))
    uvicorn.run("app:app", host="0.0.0.0", port=port, reload=False)
