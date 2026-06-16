#!/usr/bin/env python3
"""
Parse AR Aged Debtor Report Excel file and output JSON grouped by PIC.
Usage: python3 parse_ar.py <path_to_xlsx>
"""

import sys
import json
import re
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...", file=sys.stderr)
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"])
    import openpyxl


def parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y")
    return str(val)


def get_aging_bucket(row_dict):
    """Return (bucket_label, amount) for the dominant aging bucket."""
    buckets = [
        ("Over 180 ngày",  row_dict.get("over_180", 0) or 0),
        ("91-180 ngày",    row_dict.get("91_180", 0) or 0),
        ("61-90 ngày",     row_dict.get("61_90", 0) or 0),
        ("31-60 ngày",     row_dict.get("31_60", 0) or 0),
        ("1-30 ngày",      row_dict.get("1_30", 0) or 0),
        ("Current",        row_dict.get("current", 0) or 0),
    ]
    for label, amt in buckets:
        if amt and amt > 0:
            return label, amt
    return "N/A", 0


def parse_ar_file(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # Find the header row (contains 'Code' in first column)
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row[0] == 'Code':
            header_row_idx = i
            break

    if header_row_idx is None:
        raise ValueError("Không tìm thấy header row (cột 'Code')")

    # Extract report date from metadata rows
    report_date = None
    for row in ws.iter_rows(min_row=1, max_row=header_row_idx - 1, values_only=True):
        cell = str(row[0] or "")
        match = re.search(r'đến ngày[:\s]+(\d{2}/\d{2}/\d{4})', cell, re.IGNORECASE)
        if match:
            report_date = match.group(1)
            break

    # Read column headers
    headers = list(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))[0]

    col_map = {}
    for idx, h in enumerate(headers):
        if h:
            col_map[str(h).strip()] = idx

    def col(name):
        return col_map.get(name)

    # PIC is in the last column
    pic_col = len(headers) - 1

    records = []
    for row in ws.iter_rows(min_row=header_row_idx + 2, values_only=True):
        code = row[col('Code')] if col('Code') is not None else None

        if code is None or not isinstance(code, (int, float)):
            continue
        if code == 0:
            continue

        base_amount = row[col('Total of Base Amount')] if col('Total of Base Amount') is not None else None
        if base_amount is None:
            base_amount = 0

        record = {
            "code": int(code),
            "name": row[col('Customer Name')] if col('Customer Name') is not None else None,
            "transaction": row[col('Transaction Number')] if col('Transaction Number') is not None else None,
            "description": row[col('Description')] if col('Description') is not None else None,
            "invoice_no": row[col('Invoice')] if col('Invoice') is not None else None,
            "kind": row[col('Kind of Customer')] if col('Kind of Customer') is not None else None,
            "invoice_date": parse_date(row[col('Invoice Date')] if col('Invoice Date') is not None else None),
            "due_date": parse_date(row[col('Date of payment')] if col('Date of payment') is not None else None),
            "over_day": int(row[col('Over day')] or 0) if col('Over day') is not None else 0,
            "original_amount": float(row[col('Total of Original Amount')] or 0) if col('Total of Original Amount') is not None else 0,
            "base_amount": float(base_amount),
            "current":  float(row[col('Currentliability')] or 0) if col('Currentliability') is not None else 0,
            "1_30":     float(row[col('1 - 30 days')] or 0) if col('1 - 30 days') is not None else 0,
            "31_60":    float(row[col('31 - 60 days')] or 0) if col('31 - 60 days') is not None else 0,
            "61_90":    float(row[col('61 - 90 days')] or 0) if col('61 - 90 days') is not None else 0,
            "91_180":   float(row[col('91 - 180 days')] or 0) if col('91 - 180 days') is not None else 0,
            "over_180": float(row[col('Over 180 days')] or 0) if col('Over 180 days') is not None else 0,
            "type": row[col('Type')] if col('Type') is not None else None,
            "pic": row[pic_col],
        }

        bucket_label, bucket_amt = get_aging_bucket(record)
        record["aging_bucket"] = bucket_label
        record["aging_amount"] = bucket_amt

        records.append(record)

    # Group by PIC — nếu PIC trống thì dùng mã khách hàng làm key
    by_pic = {}
    for r in records:
        if r["pic"] and str(r["pic"]).strip():
            pic = str(r["pic"]).strip()
        else:
            name = str(r["name"]).strip() if r["name"] else ""
            pic = f"{r['code']}-{name}" if name else str(r["code"])
        r["pic"] = pic
        if pic not in by_pic:
            by_pic[pic] = []
        by_pic[pic].append(r)

    # Sort each PIC's invoices: oldest aging first
    aging_order = {
        "Over 180 ngày": 0, "91-180 ngày": 1, "61-90 ngày": 2,
        "31-60 ngày": 3, "1-30 ngày": 4, "Current": 5, "N/A": 6
    }
    for pic in by_pic:
        by_pic[pic].sort(key=lambda x: aging_order.get(x["aging_bucket"], 99))

    return {
        "report_date": report_date or "N/A",
        "total_records": len(records),
        "by_pic": by_pic,
    }


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python3 parse_ar.py <path_to_xlsx>", file=sys.stderr)
        sys.exit(1)

    filepath = sys.argv[1]
    result = parse_ar_file(filepath)
    sys.stdout.buffer.write(json.dumps(result, ensure_ascii=False, indent=2).encode("utf-8"))
    sys.stdout.buffer.write(b"\n")
